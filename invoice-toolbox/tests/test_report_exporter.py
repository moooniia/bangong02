import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from invoice_assistant.models import InvoiceRecord
from invoice_assistant.report_exporter import export_invoice_report


class ReportExporterTest(unittest.TestCase):
    def test_respects_selected_column_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "report.xlsx"
            record = InvoiceRecord(row_id=1, original_path="source.pdf", original_name="source.pdf", invoice_no="12345678", seller_name="测试销售方")

            export_invoice_report([record], output, field_order=["seller_name", "invoice_no"])

            sheet = load_workbook(output).active
            self.assertEqual([sheet.cell(3, 1).value, sheet.cell(3, 2).value], ["销售方名称", "发票号码"])


if __name__ == "__main__":
    unittest.main()
