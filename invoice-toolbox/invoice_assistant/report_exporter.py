from datetime import datetime
from pathlib import Path
import re
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import InvoiceRecord

FIELD_COLUMNS = [
    ("row_id", "序号"),
    ("archive_month", "归档月份"),
    ("category", "业务分类"),
    ("buyer_name", "购买方抬头"),
    ("buyer_tax", "购买方税号"),
    ("seller_name", "销售方名称"),
    ("seller_tax", "销售方税号"),
    ("invoice_date", "开票日期"),
    ("pretax_amount", "不含税金额"),
    ("tax_amount", "税额"),
    ("total_amount", "价税合计"),
    ("tax_rate", "税率"),
    ("invoice_type", "发票类型"),
    ("invoice_no", "发票号码"),
    ("line_items", "商品/服务"),
    ("original_name", "原文件名"),
    ("archived_name", "新文件名"),
    ("original_path", "原文件链接"),
    ("archived_path", "归档文件链接"),
    ("status", "审核状态"),
    ("review_summary", "审核原因"),
]

REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
ALT_FILL = PatternFill(fill_type="solid", fgColor="F5FBF7")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="C8EDD9")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1A6B4A")
REVIEW_FONT = Font(name="Calibri", size=11, bold=True, color="C00000")
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
NORMAL_FONT = Font(name="Calibri", size=11)
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1A6B4A")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="888888")
TOTAL_FONT = Font(name="微软雅黑", size=12, bold=True, color="1A6B4A")
TOTAL_NUM_FONT = Font(name="Consolas", size=12, bold=True, color="1A6B4A")
AMOUNT_FONT = Font(name="Consolas", size=10)


def export_invoice_report(records: Iterable[InvoiceRecord], output_path: Path, field_order=None) -> Path:
    rows = list(records)
    field_map = dict(FIELD_COLUMNS)
    columns = [(field, field_map[field]) for field in (field_order or [item[0] for item in FIELD_COLUMNS]) if field in field_map]
    if not columns:
        columns = [("invoice_no", field_map["invoice_no"])]
    wb = Workbook()
    ws = wb.active
    ws.title = "发票明细"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(1, 1, "发票整理报表").font = TITLE_FONT
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(2, 1, f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}    共 {len(rows)} 张")
    ws.cell(2, 1).alignment = center
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 18

    for column_index, (_, title) in enumerate(columns, 1):
        cell = ws.cell(3, column_index, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 20

    for row_index, record in enumerate(rows, 4):
        record_dict = record.as_dict()
        for column_index, (field_name, _) in enumerate(columns, 1):
            value = _clean_excel_value(_value_for_field(record, field_name, record_dict))
            is_review_field = field_name in record.fields_needing_review
            if is_review_field and (value == "" or value is None):
                value = "待确认"
            cell = ws.cell(row_index, column_index, value)
            cell.border = border
            cell.font = NORMAL_FONT
            cell.alignment = right if field_name in {"pretax_amount", "tax_amount", "total_amount"} else left
            if row_index % 2 == 0:
                cell.fill = ALT_FILL
            if field_name in {"pretax_amount", "tax_amount", "total_amount"} and value:
                cell.font = AMOUNT_FONT
                cell.number_format = "#,##0.00"
            if is_review_field:
                cell.fill = REVIEW_FILL
                if value == "待确认":
                    cell.font = REVIEW_FONT
            if field_name in {"original_path", "archived_path"} and value:
                cell.value = "📎 打开文件"
                cell.hyperlink = value
                cell.font = LINK_FONT

    _write_total_row(ws, rows, border, right, columns)
    _fit_dimensions(ws)

    ws.freeze_panes = "A4"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_total_row(ws, rows: List[InvoiceRecord], border: Border, right: Alignment, columns) -> None:
    row_index = len(rows) + 4
    total_column = next((index for index, (field, _) in enumerate(columns, 1) if field == "total_amount"), None)
    label_end = max(1, (total_column or len(columns) + 1) - 1)
    if label_end > 1:
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=label_end)
    label = ws.cell(row_index, 1, f"合  计（共 {len(rows)} 张）")
    label.font = TOTAL_FONT
    label.fill = TOTAL_FILL
    label.alignment = right
    label.border = border
    for column_index in range(2, label_end + 1):
        cell = ws.cell(row_index, column_index)
        cell.fill = TOTAL_FILL
        cell.border = border

    if total_column:
        total_cell = ws.cell(row_index, total_column, _sum_amounts(rows))
        total_cell.font = TOTAL_NUM_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.alignment = right
        total_cell.border = border
        total_cell.number_format = "#,##0.00"
    for column_index in range(label_end + 1, len(columns) + 1):
        if column_index == total_column:
            continue
        cell = ws.cell(row_index, column_index)
        cell.fill = TOTAL_FILL
        cell.border = border
    ws.row_dimensions[row_index].height = 24


def _sum_amounts(rows: List[InvoiceRecord]) -> float:
    total = 0.0
    for record in rows:
        try:
            total += float(str(record.total_amount).replace(",", ""))
        except (TypeError, ValueError):
            continue
    return round(total, 2)


def _clean_excel_value(value):
    if not isinstance(value, str):
        return value
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)


def _fit_dimensions(ws) -> None:
    min_widths = {
        1: 8,
        2: 12,
        3: 12,
        8: 12,
        9: 12,
        10: 10,
        11: 12,
        12: 8,
        18: 14,
        19: 14,
    }
    max_widths = {
        4: 34,
        5: 24,
        6: 34,
        7: 24,
        14: 24,
        15: 34,
        16: 50,
        17: 58,
        18: 16,
        19: 16,
    }
    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        max_len = 0
        for row_index in range(1, ws.max_row + 1):
            if row_index in (1, 2):
                continue
            if row_index == ws.max_row and column_index <= 10:
                continue
            value = ws.cell(row_index, column_index).value
            if value is None:
                continue
            max_len = max(max_len, _display_width(str(value)))
        width = max_len + 2
        width = max(width, min_widths.get(column_index, 8))
        width = min(width, max_widths.get(column_index, 40))
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 18
    for row_index in range(4, ws.max_row):
        ws.row_dimensions[row_index].height = 18


def _display_width(value: str) -> int:
    width = 0
    for char in value:
        width += 2 if ord(char) > 127 else 1
    return width


def records_from_dicts(items: List[dict]) -> List[InvoiceRecord]:
    return [InvoiceRecord.from_dict(item) for item in items]


def _value_for_field(record: InvoiceRecord, field_name: str, record_dict: dict) -> str:
    if field_name == "row_id":
        return record.row_id
    if field_name == "archive_month":
        if record.invoice_date and len(record.invoice_date) >= 7:
            year, month = record.invoice_date[:7].split("-")
            return f"{year}年{month}月"
        return ""
    if field_name == "archived_name":
        return Path(record.archived_path).name if record.archived_path else ""
    if field_name == "review_summary":
        if not record.fields_needing_review:
            return ""
        parts = []
        for field in sorted(record.fields_needing_review):
            reasons = record.review_reasons.get(field, [])
            parts.append(f"{field}: {'；'.join(reasons) if reasons else '待确认'}")
        return " | ".join(parts)
    value = record_dict.get(field_name, "")
    if field_name in {"pretax_amount", "tax_amount", "total_amount"} and value not in ("", None):
        try:
            return float(str(value).replace(",", ""))
        except ValueError:
            return str(value)
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return "" if value is None else str(value)
